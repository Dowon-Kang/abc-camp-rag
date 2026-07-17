import re
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
CSV_PATH = os.path.join(DATA_DIR, "yes24_it_mobile_bestseller.csv")
OUTPUT_PATH = os.path.join(DATA_DIR, "yes24_it_mobile_bestseller_dashboard.xlsx")


def clean_price(val):
    if pd.isna(val) or val == "":
        return 0
    return int(re.sub(r"[^\d]", "", str(val)))


def extract_year(val):
    if pd.isna(val) or val == "":
        return None
    m = re.search(r"(\d{4})", str(val))
    return int(m.group(1)) if m else None


def extract_ym(val):
    if pd.isna(val) or val == "":
        return None
    m = re.search(r"(\d{4})년\s*(\d{1,2})월", str(val))
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else None


def load_data():
    df = pd.read_csv(CSV_PATH, encoding="utf-8-sig")
    df["판매가_num"] = df["판매가"].apply(clean_price)
    df["정가_num"] = df["정가"].apply(clean_price)
    df["할인율_num"] = pd.to_numeric(df["할인율"], errors="coerce").fillna(0).astype(int)
    df["출간년도"] = df["출간일"].apply(extract_year)
    df["출간년월"] = df["출간일"].apply(extract_ym)
    return df


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="맑은 고딕", size=12, bold=True, color="2E75B6")
KPI_VALUE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1F4E79")
KPI_LABEL_FONT = Font(name="맑은 고딕", size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except:
                pass
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def create_summary_sheet(wb, df):
    ws = wb.active
    ws.title = "대시보드 요약"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Yes24 IT 모바일 종합 베스트셀러 대시보드"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"총 {len(df):,}권 | {df['출판사'].nunique()}개 출판사 | 가격 범위 {df['판매가_num'].min():,}~{df['판매가_num'].max():,}원"
    ws["A2"].font = Font(name="맑은 고딕", size=10, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    kpi_data = [
        ("총 도서 수", f"{len(df):,}권", "A"),
        ("평균 판매가", f"{df['판매가_num'].mean():,.0f}원", "C"),
        ("평균 할인율", f"{df['할인율_num'].mean():.1f}%", "E"),
        ("출판사 수", f"{df['출판사'].nunique():,}곳", "G"),
    ]
    for label, value, col_letter in kpi_data:
        label_cell = ws[f"{col_letter}3"]
        label_cell.value = label
        label_cell.font = KPI_LABEL_FONT
        label_cell.alignment = Alignment(horizontal="center")
        val_cell = ws[f"{col_letter}4"]
        val_cell.value = value
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{col_letter}3:{col_letter}4")

    for col in range(1, 9):
        ws.cell(row=4, column=col).border = THIN_BORDER

    ws["A6"] = "출판사별 도서 수 (상위 15)"
    ws["A6"].font = SUBTITLE_FONT

    pub_counts = df["출판사"].value_counts().head(15).reset_index()
    pub_counts.columns = ["출판사", "도서 수"]
    start_row = 8
    ws.cell(row=start_row, column=1, value="출판사")
    ws.cell(row=start_row, column=2, value="도서 수")
    style_header_row(ws, start_row, 2)
    for i, (_, row) in enumerate(pub_counts.iterrows()):
        ws.cell(row=start_row + 1 + i, column=1, value=row["출판사"]).border = THIN_BORDER
        ws.cell(row=start_row + 1 + i, column=2, value=row["도서 수"]).border = THIN_BORDER
        ws.cell(row=start_row + 1 + i, column=2).alignment = Alignment(horizontal="center")

    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "출판사별 도서 수"
    chart1.y_axis.title = "출판사"
    chart1.x_axis.title = "도서 수"
    chart1.width = 22
    chart1.height = 14
    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 15)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 15)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, "D8")

    ws["A25"] = "출간년도별 도서 수"
    ws["A25"].font = SUBTITLE_FONT

    year_counts = df["출간년도"].dropna().astype(int).value_counts().sort_index().reset_index()
    year_counts.columns = ["출간년도", "도서 수"]
    yr = 26
    ws.cell(row=yr, column=1, value="출간년도")
    ws.cell(row=yr, column=2, value="도서 수")
    style_header_row(ws, yr, 2)
    for i, (_, row) in enumerate(year_counts.iterrows()):
        ws.cell(row=yr + 1 + i, column=1, value=int(row["출간년도"])).border = THIN_BORDER
        ws.cell(row=yr + 1 + i, column=2, value=row["도서 수"]).border = THIN_BORDER
        ws.cell(row=yr + 1 + i, column=2).alignment = Alignment(horizontal="center")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "출간년도별 도서 수"
    chart2.y_axis.title = "도서 수"
    chart2.x_axis.title = "출간년도"
    chart2.width = 22
    chart2.height = 14
    data2 = Reference(ws, min_col=2, min_row=yr, max_row=yr + len(year_counts))
    cats2 = Reference(ws, min_col=1, min_row=yr + 1, max_row=yr + len(year_counts))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "D25")

    auto_width(ws, min_width=12, max_width=30)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12


def create_price_sheet(wb, df):
    ws = wb.create_sheet("가격 분석")
    ws.sheet_properties.tabColor = "2E75B6"

    ws.merge_cells("A1:F1")
    ws["A1"] = "가격 분석"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    ws["A3"] = "가격 구간"
    ws["B3"] = "도서 수"
    ws["C3"] = "평균 할인율"
    ws["D3"] = "평균 판매가"
    style_header_row(ws, 3, 4)

    bins = [
        ("~10,000원", 0, 10000),
        ("10,001~20,000원", 10001, 20000),
        ("20,001~30,000원", 20001, 30000),
        ("30,001~40,000원", 30001, 40000),
        ("40,001~50,000원", 40001, 50000),
        ("50,001원~", 50001, 999999),
    ]
    for i, (label, lo, hi) in enumerate(bins):
        subset = df[(df["판매가_num"] >= lo) & (df["판매가_num"] <= hi)]
        r = 4 + i
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=len(subset)).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=round(subset["할인율_num"].mean(), 1) if len(subset) > 0 else 0).border = THIN_BORDER
        ws.cell(row=r, column=3).number_format = "0.0"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=round(subset["판매가_num"].mean()) if len(subset) > 0 else 0).border = THIN_BORDER
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

    chart_price = BarChart()
    chart_price.type = "col"
    chart_price.style = 10
    chart_price.title = "가격 구간별 도서 수"
    chart_price.y_axis.title = "도서 수"
    chart_price.width = 20
    chart_price.height = 12
    data = Reference(ws, min_col=2, min_row=3, max_row=9)
    cats = Reference(ws, min_col=1, min_row=4, max_row=9)
    chart_price.add_data(data, titles_from_data=True)
    chart_price.set_categories(cats)
    ws.add_chart(chart_price, "A12")

    ws.merge_cells("A29:F29")
    ws["A29"] = "할인율 구간별 분석"
    ws["A29"].font = SUBTITLE_FONT

    ws["A30"] = "할인율 구간"
    ws["B30"] = "도서 수"
    ws["C30"] = "평균 판매가"
    style_header_row(ws, 30, 3)

    disc_bins = [
        ("0% (할인 없음)", 0, 0),
        ("1~5%", 1, 5),
        ("6~10%", 6, 10),
        ("11~15%", 11, 15),
        ("16~20%", 16, 20),
        ("21%~", 21, 100),
    ]
    for i, (label, lo, hi) in enumerate(disc_bins):
        subset = df[(df["할인율_num"] >= lo) & (df["할인율_num"] <= hi)]
        r = 31 + i
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=len(subset)).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=round(subset["판매가_num"].mean()) if len(subset) > 0 else 0).border = THIN_BORDER
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")

    chart_disc = BarChart()
    chart_disc.type = "col"
    chart_disc.style = 10
    chart_disc.title = "할인율 구간별 도서 수"
    chart_disc.y_axis.title = "도서 수"
    chart_disc.width = 20
    chart_disc.height = 12
    data_d = Reference(ws, min_col=2, min_row=30, max_row=36)
    cats_d = Reference(ws, min_col=1, min_row=31, max_row=36)
    chart_disc.add_data(data_d, titles_from_data=True)
    chart_disc.set_categories(cats_d)
    ws.add_chart(chart_disc, "A38")

    top20 = df.nlargest(20, "판매가_num")[["순위", "제목", "저자", "출판사", "판매가", "정가", "할인율"]]
    ws["A56"] = "상위 20권 (가격 기준)"
    ws["A56"].font = SUBTITLE_FONT
    headers = list(top20.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=57, column=c, value=h)
    style_header_row(ws, 57, len(headers))
    for i, (_, row) in enumerate(top20.iterrows()):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=58 + i, column=c, value=row[h])
            cell.border = THIN_BORDER
            if h in ("판매가", "정가"):
                cell.alignment = Alignment(horizontal="right")

    auto_width(ws, min_width=12, max_width=50)
    ws.column_dimensions["A"].width = 22


def create_publisher_sheet(wb, df):
    ws = wb.create_sheet("출판사 분석")
    ws.sheet_properties.tabColor = "70AD47"

    ws.merge_cells("A1:F1")
    ws["A1"] = "출판사별 상세 분석"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    pub_stats = (
        df.groupby("출판사")
        .agg(
            도서_수=("제목", "count"),
            평균_판매가=("판매가_num", "mean"),
            평균_할인율=("할인율_num", "mean"),
            최고_판매가=("판매가_num", "max"),
            최저_판매가=("판매가_num", "min"),
        )
        .sort_values("도서_수", ascending=False)
        .reset_index()
    )
    pub_stats["평균_판매가"] = pub_stats["평균_판매가"].round(0).astype(int)
    pub_stats["평균_할인율"] = pub_stats["평균_할인율"].round(1)

    headers = list(pub_stats.columns)
    h_map = {
        "출판사": "출판사", "도서_수": "도서 수", "평균_판매가": "평균 판매가",
        "평균_할인율": "평균 할인율(%)", "최고_판매가": "최고 판매가", "최저_판매가": "최저 판매가",
    }
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h_map.get(h, h))
    style_header_row(ws, 3, len(headers))

    for i, (_, row) in enumerate(pub_stats.iterrows()):
        r = 4 + i
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=row[h])
            cell.border = THIN_BORDER
            if h in ("평균_판매가", "최고_판매가", "최저_판매가"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif h == "도서_수":
                cell.alignment = Alignment(horizontal="center")
            elif h == "평균_할인율":
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")

    chart_pub = BarChart()
    chart_pub.type = "bar"
    chart_pub.style = 10
    chart_pub.title = "출판사별 도서 수 (상위 20)"
    chart_pub.y_axis.title = "출판사"
    chart_pub.width = 22
    chart_pub.height = 16
    end_row = min(3 + 20, 3 + len(pub_stats))
    data = Reference(ws, min_col=2, min_row=3, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    chart_pub.add_data(data, titles_from_data=True)
    chart_pub.set_categories(cats)
    ws.add_chart(chart_pub, "H3")

    chart_avg = BarChart()
    chart_avg.type = "bar"
    chart_avg.style = 10
    chart_avg.title = "출판사별 평균 판매가 (상위 20)"
    chart_avg.y_axis.title = "출판사"
    chart_avg.width = 22
    chart_avg.height = 16
    data2 = Reference(ws, min_col=3, min_row=3, max_row=end_row)
    cats2 = Reference(ws, min_col=1, min_row=4, max_row=end_row)
    chart_avg.add_data(data2, titles_from_data=True)
    chart_avg.set_categories(cats2)
    ws.add_chart(chart_avg, "H21")

    auto_width(ws, min_width=14, max_width=30)
    ws.column_dimensions["A"].width = 28


def create_data_sheet(wb, df):
    ws = wb.create_sheet("전체 데이터")
    ws.sheet_properties.tabColor = "FFC000"

    display_cols = ["순위", "제목", "저자", "출판사", "출간일", "판매가", "정가", "할인율", "링크"]
    headers_map = {
        "순위": "순위", "제목": "제목", "저자": "저자", "출판사": "출판사",
        "출간일": "출간일", "판매가": "판매가", "정가": "정가", "할인율": "할인율(%)", "링크": "링크",
    }

    for c, col in enumerate(display_cols, 1):
        ws.cell(row=1, column=c, value=headers_map.get(col, col))
    style_header_row(ws, 1, len(display_cols))

    for i, (_, row) in enumerate(df.iterrows()):
        r = 2 + i
        for c, col in enumerate(display_cols, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
            if col == "순위":
                cell.alignment = Alignment(horizontal="center")
            elif col in ("판매가", "정가"):
                cell.alignment = Alignment(horizontal="right")
            elif col == "할인율":
                cell.alignment = Alignment(horizontal="center")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(display_cols))}{len(df) + 1}"
    ws.freeze_panes = "A2"

    auto_width(ws, min_width=10, max_width=60)
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["I"].width = 45


def create_ranking_sheet(wb, df):
    ws = wb.create_sheet("Top 50 순위")
    ws.sheet_properties.tabColor = "ED7D31"

    ws.merge_cells("A1:H1")
    ws["A1"] = "베스트셀러 Top 50"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    top50 = df.head(50)[["순위", "제목", "저자", "출판사", "출간일", "판매가", "정가", "할인율"]]
    headers = ["순위", "제목", "저자", "출판사", "출간일", "판매가", "정가", "할인율(%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    top_fills = [gold_fill, silver_fill, bronze_fill]

    for i, (_, row) in enumerate(top50.iterrows()):
        r = 4 + i
        for c, col in enumerate(["순위", "제목", "저자", "출판사", "출간일", "판매가", "정가", "할인율"], 1):
            cell = ws.cell(row=r, column=c, value=row[col])
            cell.border = THIN_BORDER
            if col == "순위":
                cell.alignment = Alignment(horizontal="center")
            elif col in ("판매가", "정가"):
                cell.alignment = Alignment(horizontal="right")
            elif col == "할인율":
                cell.alignment = Alignment(horizontal="center")
        if i < 3:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = top_fills[i]
                ws.cell(row=r, column=c).font = Font(bold=True)

    chart_top = BarChart()
    chart_top.type = "col"
    chart_top.style = 10
    chart_top.title = "Top 20 판매가 비교"
    chart_top.y_axis.title = "판매가 (원)"
    chart_top.width = 24
    chart_top.height = 14
    data = Reference(ws, min_col=6, min_row=3, max_row=23)
    cats = Reference(ws, min_col=1, min_row=4, max_row=23)
    chart_top.add_data(data, titles_from_data=True)
    chart_top.set_categories(cats)
    ws.add_chart(chart_top, "J3")

    auto_width(ws, min_width=10, max_width=55)
    ws.column_dimensions["B"].width = 50


def create_monthly_sheet(wb, df):
    ws = wb.create_sheet("월별 추이")
    ws.sheet_properties.tabColor = "5B9BD5"

    ws.merge_cells("A1:F1")
    ws["A1"] = "출간 월별 도서 수 추이"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    monthly = df["출간년월"].dropna().value_counts().sort_index().reset_index()
    monthly.columns = ["출간년월", "도서 수"]

    ws.cell(row=3, column=1, value="출간년월")
    ws.cell(row=3, column=2, value="도서 수")
    style_header_row(ws, 3, 2)

    for i, (_, row) in enumerate(monthly.iterrows()):
        r = 4 + i
        ws.cell(row=r, column=1, value=row["출간년월"]).border = THIN_BORDER
        ws.cell(row=r, column=2, value=row["도서 수"]).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    chart_monthly = BarChart()
    chart_monthly.type = "col"
    chart_monthly.style = 10
    chart_monthly.title = "출간 월별 도서 수"
    chart_monthly.y_axis.title = "도서 수"
    chart_monthly.x_axis.title = "출간월"
    chart_monthly.width = 28
    chart_monthly.height = 14
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(monthly))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(monthly))
    chart_monthly.add_data(data, titles_from_data=True)
    chart_monthly.set_categories(cats)
    ws.add_chart(chart_monthly, "D3")

    auto_width(ws, min_width=14, max_width=20)


def main():
    df = load_data()
    wb = Workbook()

    create_summary_sheet(wb, df)
    create_price_sheet(wb, df)
    create_publisher_sheet(wb, df)
    create_ranking_sheet(wb, df)
    create_monthly_sheet(wb, df)
    create_data_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"[DONE] 대시보드 생성 완료: {OUTPUT_PATH}")
    print(f"  - 총 {len(df):,}권 데이터 포함")
    print(f"  - 6개 시트: 대시보드 요약, 가격 분석, 출판사 분석, Top 50 순위, 월별 추이, 전체 데이터")


if __name__ == "__main__":
    main()
